import pandas as pd
import numpy as np
import calendar
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

BASE = 'Perform AI Team'
FONT_NAME = 'Arial'

# =====================================================================
# 1. LOAD & PREP DATA (python side — classification / parsing only)
# =====================================================================
master = pd.read_excel(f'{BASE}/Master_รายชื่อพนักงานปัจจุบัน-PDE.xlsx')
master['full_name'] = (master['ชื่อ'].str.strip() + ' ' + master['สกุล'].str.strip()).str.replace(r'\s+', ' ', regex=True)

mapping = pd.read_excel(f'{BASE}/JobRole_Mapping_PCC_Database_R.0.xlsx', sheet_name='Mapping Result')
pde_map = mapping[mapping['บริษัท (Code)'] == 'PDE'].copy()

# --- Join key changed per user request: link via Job Role Code (เดิม) <-> Master's
# Job Role Code, NOT via employee code. Multiple PDE employees can share the same
# Job Role Code; where the mapping file has more than one row for that code (different
# employees individually assessed), keep the row with the highest match score and
# flag the code as a conflict for visibility. This also recovers cases where one
# employee's own mapping row has a blank Job Code (PCC Database) but a colleague
# holding the identical Job Role Code has it filled in.
job_role_code_conflicts = sorted(
    code for code, n in pde_map.groupby('Job Role Code (เดิม)')['Job Code (PCC Database)'].nunique().items() if n > 1
)
pde_map_sorted = pde_map.sort_values('คะแนนความใกล้เคียง (%)', ascending=False)
job_role_lookup = pde_map_sorted.drop_duplicates(subset='Job Role Code (เดิม)', keep='first').copy()
job_role_lookup['conflict_flag'] = job_role_lookup['Job Role Code (เดิม)'].isin(job_role_code_conflicts)

master_full = master.merge(
    job_role_lookup[['Job Role Code (เดิม)', 'Job Code (PCC Database)', 'ชื่อตำแหน่งใน Job Role Database',
                      'ชื่อตำแหน่ง (TH)', 'Job Family', 'Pipeline Level', 'Job Level ของ Pipeline',
                      'ผลรวม JE', 'คะแนนความใกล้เคียง (%)', 'สถานะการแมส', 'conflict_flag']],
    left_on='Job Role Code', right_on='Job Role Code (เดิม)', how='left'
)
master_full['conflict_flag'] = master_full['conflict_flag'].fillna(False)

ts = pd.read_csv('all_timesheet.csv')
ts['employee_name'] = ts['employee_name'].str.replace(r'\s+', ' ', regex=True).str.strip()
ts = ts[ts['employee_name'] != 'TEMP'].copy()


def categorize(job_no, desc):
    j = str(job_no).upper()
    d = str(desc)
    if 'LEAVE' in j or 'ลาป่วย' in d or 'ลากิจ' in d or 'ลาพักร้อน' in d:
        return 'Leave / ลา'
    if 'ADMIN' in j:
        return 'Admin / งานธุรการ'
    if re.search(r'IT SUPPORT', d.upper()):
        return 'IT Support (ticket)'
    if j.startswith('PJ-') or j.startswith('DBA-') or j.startswith('MA-') or 'GRP-RPM' in j or 'ICT-' in j:
        return 'Project work'
    if re.match(r'^\d{3}-', j):
        return 'Project / Job code'
    return 'Other / ไม่ระบุหมวด'


ts['category'] = ts.apply(lambda r: categorize(r['job_no'], r['description']), axis=1)

MONTH_ORDER = [(2026, 1, 'Jan'), (2026, 2, 'Feb'), (2026, 3, 'Mar'), (2026, 4, 'April'),
               (2026, 5, 'May'), (2026, 6, 'June'), (2026, 7, 'July')]


def std_hours(year, month):
    days_in_month = calendar.monthrange(year, month)[1]
    weekdays = sum(1 for d in range(1, days_in_month + 1) if calendar.weekday(year, month, d) < 5)
    return weekdays * 8


std_table = pd.DataFrame([{'year': y, 'month': m, 'month_name': mn, 'std_hours': std_hours(y, m)} for y, m, mn in MONTH_ORDER])

# distinct job codes per employee (python — hard to do reliably as a live formula)
distinct_jobs = ts.groupby('employee_name')['job_no'].nunique().rename('distinct_jobs')
months_active = ts.groupby('employee_name')['month'].nunique().rename('months_active')
headcount_per_month = ts.groupby(['year', 'month'])['employee_name'].nunique().rename('headcount').reset_index()

# union of names: master + timesheet-only
ts_names = set(ts['employee_name'].unique())
m_names = set(master_full['full_name'].unique())
ts_only = sorted(ts_names - m_names)

emp_rows = []
for _, r in master_full.iterrows():
    emp_rows.append({
        'full_name': r['full_name'], 'Employee Code': r['Employee Code'], 'Job Role': r['Job Role'],
        'Job Role Code': r['Job Role Code'],
        'Position': r['Position'], 'Employee Level': r['Employee Level'],
        'Department': r['ชื่อหน่วยงาน (ระดับ 4)'], 'Job Code (PCC)': r['Job Code (PCC Database)'],
        'Job Role DB (TH)': r['ชื่อตำแหน่ง (TH)'], 'Job Family': r['Job Family'],
        'JE รวม': r['ผลรวม JE'], 'Match %': r['คะแนนความใกล้เคียง (%)'], 'สถานะการแมส': r['สถานะการแมส'],
        'มี Job Role Code ซ้ำกับพนักงานอื่นที่แมสต่างกัน': 'ใช่ - ควรตรวจสอบ' if r['conflict_flag'] else '',
        'in_master': 'Y', 'in_timesheet': 'Y' if r['full_name'] in ts_names else 'N',
    })
for name in ts_only:
    emp_rows.append({
        'full_name': name, 'Employee Code': None, 'Job Role': None, 'Job Role Code': None, 'Position': None, 'Employee Level': None,
        'Department': None, 'Job Code (PCC)': None, 'Job Role DB (TH)': None, 'Job Family': None,
        'JE รวม': None, 'Match %': None, 'สถานะการแมส': 'ไม่พบใน Master (อาจเข้าใหม่/พ้นสภาพ)',
        'มี Job Role Code ซ้ำกับพนักงานอื่นที่แมสต่างกัน': '',
        'in_master': 'N', 'in_timesheet': 'Y',
    })
emp_master_df = pd.DataFrame(emp_rows)

print('rows raw ts:', len(ts))
print('employees union:', len(emp_master_df))
print(emp_master_df['สถานะการแมส'].value_counts())

# =====================================================================
# 2. BUILD WORKBOOK
# =====================================================================
wb = Workbook()

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color='1F4E79')
SUB_FONT = Font(name=FONT_NAME, size=11, color='52514E')
BODY_FONT = Font(name=FONT_NAME, size=10.5)
BOLD_FONT = Font(name=FONT_NAME, size=10.5, bold=True)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND_FILL = PatternFill('solid', fgColor='F2F2F2')


def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_df(ws, df, start_row, start_col=1, number_formats=None, band=True):
    number_formats = number_formats or {}
    ncols = len(df.columns)
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=col)
    style_header_row(ws, start_row, ncols, start_col)
    for i, (_, row) in enumerate(df.iterrows()):
        r = start_row + 1 + i
        for j, col in enumerate(df.columns):
            val = row[col]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=r, column=start_col + j, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if band and i % 2 == 1:
                cell.fill = BAND_FILL
            if col in number_formats:
                cell.number_format = number_formats[col]
    return start_row + 1 + len(df)


# ---------------------------------------------------------------
# Sheet 1: วิธีใช้งาน
# ---------------------------------------------------------------
ws = wb.active
ws.title = 'วิธีใช้งาน'
ws.sheet_view.showGridLines = False
ws['B2'] = 'รายงานวิเคราะห์ข้อมูลพนักงาน PDE'
ws['B2'].font = TITLE_FONT
ws['B3'] = 'จาก Job Role, Daily Report (Timesheet) และ Master รายชื่อพนักงาน — มกราคม–กรกฎาคม 2569'
ws['B3'].font = SUB_FONT

lines = [
    ('แหล่งข้อมูล', ''),
    ('', '1. Master_รายชื่อพนักงานปัจจุบัน-PDE.xlsx — รายชื่อพนักงาน, Job Role, Job Role Code, หน่วยงาน (42 คน)'),
    ('', '2. JobRole_Mapping_PCC_Database_R.0.xlsx — ผลการจับคู่ Job Role Code (เดิม) กับ Job Code (PCC Database) พร้อมคะแนน JE และ % ความใกล้เคียง'),
    ('', '3. Timesheet PDE_[เดือน] 2026.xls — Daily Report รายวันจากระบบ Timesheet (ม.ค.–ก.ค. 2569 รวม 7 ไฟล์)'),
    ('', ''),
    ('วิธีเชื่อมข้อมูล Job Role Mapping (ปรับตามที่ผู้ใช้ระบุ)', ''),
    ('', 'เชื่อม Master (คอลัมน์ "Job Role Code") กับไฟล์ Mapping (คอลัมน์ "Job Role Code (เดิม)") โดยตรง — ไม่ใช้รหัสพนักงานเป็นตัวเชื่อมอีกต่อไป'),
    ('', 'เหตุผล: Job Code (PCC Database) เป็นคุณสมบัติของ "ตำแหน่ง/Job Role" ไม่ใช่ของพนักงานรายคน พนักงานที่ถือ Job Role Code เดียวกันควรได้ Job Code (PCC) เดียวกัน'),
    ('', 'บาง Job Role Code มีมากกว่า 1 แถวในไฟล์ Mapping (ประเมินแยกรายคน) และบางแถวมีผลต่างกัน — ระบบเลือกแถวที่มีคะแนนความใกล้เคียง (%) สูงสุดเป็นค่าอ้างอิงของ Job Role Code นั้น และทำเครื่องหมาย "ใช่ - ควรตรวจสอบ" ไว้ในชีต Master พนักงาน + Job Role'),
    ('', 'พบ Job Role Code ที่มีผลแมสต่างกัน 2 รหัส จากทั้งหมด 24 รหัสที่ใช้ในกลุ่ม PDE (ดูคอลัมน์ "มี Job Role Code ซ้ำ...") ควรให้ผู้รับผิดชอบ Job Role Database ยืนยันผลที่ถูกต้อง'),
    ('', ''),
    ('เนื้อหาแต่ละ Sheet', ''),
    ('', 'สรุปภาพรวม — แนวโน้มชั่วโมงทำงานและอัตราการบันทึก Timesheet รายเดือน'),
    ('', 'สรุปรายแผนก — ชั่วโมงทำงานรวมและจำนวนพนักงานแยกตามฝ่าย'),
    ('', 'Job Role Mapping Quality — สถานะการจับคู่ Job Role ของพนักงาน PDE กับฐานข้อมูลกลาง'),
    ('', 'สรุปรายบุคคล — ชั่วโมงทำงานรวมและสัดส่วนกิจกรรม (Project / Admin / Leave / IT Support) ต่อคน'),
    ('', 'Master พนักงาน + Job Role — ตารางอ้างอิงพนักงานที่รวม Job Role และผลคะแนนการจับคู่'),
    ('', 'ข้อมูลดิบ Timesheet — รายการบันทึกเวลาทำงานทุกแถว พร้อมคอลัมน์หมวดกิจกรรมที่จัดกลุ่มให้'),
    ('', ''),
    ('วิธีจัดหมวดกิจกรรม (คอลัมน์ "หมวดกิจกรรม" ในข้อมูลดิบ)', ''),
    ('', 'จัดกลุ่มจาก Job No. / Description ด้วยกฎคำสำคัญ (Leave, Admin, IT Support ticket, Project) — เป็นการจัดหมวดเบื้องต้น'),
    ('', 'ยังไม่ใช่การเทียบ Activity กับ Job Role โดยตรง เพราะ Job No./Description ไม่ได้ผูกรหัสกับ Job Role ในระบบต้นทาง'),
    ('', 'เป้าหมายขั้นถัดไปคือสร้างตาราง Mapping ระหว่าง Job No./โครงการ กับ Job Role ตามที่เสนอไว้ในข้อเสนอโครงการ (Phase 1)'),
    ('', ''),
    ('ข้อควรระวังของข้อมูล', ''),
    ('', 'ชั่วโมงมาตรฐาน/เดือน คำนวณจาก (จำนวนวันทำการ จ.-ศ. ในเดือน) x 8 ชม. เป็นค่าประมาณ ยังไม่หักวันหยุดนักขัตฤกษ์'),
    ('', 'พนักงาน 4 คนในไฟล์ Master ไม่มีข้อมูล Timesheet ในช่วงนี้ (อาจเพิ่งเข้าใหม่/ยังไม่บันทึก)'),
    ('', 'พนักงาน 7 คนมีข้อมูล Timesheet แต่ไม่พบในไฟล์ Master (อาจพ้นสภาพระหว่างช่วง หรือชื่อสะกดไม่ตรงกัน) — ตรวจสอบเพิ่มเติมกับ HR'),
    ('', 'ตัดรายการที่ชื่อพนักงานเป็น "TEMP" ออกจากการวิเคราะห์ทั้งหมด'),
]
r = 5
for label, text in lines:
    if label:
        cell = ws.cell(row=r, column=2, value=label)
        cell.font = BOLD_FONT
    else:
        cell = ws.cell(row=r, column=2, value=('   • ' + text) if text else '')
        cell.font = BODY_FONT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    r += 1
autofit(ws, [3] + [14] * 9)

# ---------------------------------------------------------------
# Sheet 2: สรุปภาพรวม (trend, with SUMIFS/COUNTIFS formulas -> RawData)
# ---------------------------------------------------------------
ws2 = wb.create_sheet('สรุปภาพรวมรายเดือน')
ws2.sheet_view.showGridLines = False
ws2['B2'] = 'แนวโน้มชั่วโมงทำงานรายเดือน (จาก Timesheet)'
ws2['B2'].font = TITLE_FONT
raw_sheet_name = 'ข้อมูลดิบ Timesheet'

headers2 = ['เดือน', 'ปี', 'จำนวนพนักงานที่บันทึก (Headcount)', 'ชั่วโมงมาตรฐาน/คน', 'ชั่วโมงมาตรฐานรวม',
            'ชั่วโมงที่บันทึกจริง', 'อัตราการบันทึก (%)']
start_r = 4
for j, h in enumerate(headers2):
    ws2.cell(row=start_r, column=2 + j, value=h)
style_header_row(ws2, start_r, len(headers2), start_col=2)

n_raw = len(ts)
raw_last_row = n_raw + 1  # header at row1 in RawData

for i, (y, m, mn) in enumerate(MONTH_ORDER):
    r = start_r + 1 + i
    hc = int(headcount_per_month[(headcount_per_month.year == y) & (headcount_per_month.month == m)]['headcount'].iloc[0])
    sh = std_hours(y, m)
    ws2.cell(row=r, column=2, value=mn).font = BODY_FONT
    ws2.cell(row=r, column=3, value=y).font = BODY_FONT
    ws2.cell(row=r, column=4, value=hc).font = BODY_FONT  # precomputed distinct headcount (documented below)
    ws2.cell(row=r, column=5, value=sh).font = BODY_FONT
    ws2.cell(row=r, column=6, value=f'=D{r}*E{r}').font = BODY_FONT
    ws2.cell(row=r, column=7, value=(
        f"=SUMIFS('{raw_sheet_name}'!$H$2:$H${raw_last_row},"
        f"'{raw_sheet_name}'!$B$2:$B${raw_last_row},C{r},"
        f"'{raw_sheet_name}'!$C$2:$C${raw_last_row},MATCH(B{r},{{\"Jan\",\"Feb\",\"Mar\",\"April\",\"May\",\"June\",\"July\"}},0))"
    )).font = BODY_FONT
    ws2.cell(row=r, column=8, value=f'=IFERROR(G{r}/F{r},0)').font = BODY_FONT
    ws2.cell(row=r, column=8).number_format = '0.0%'
    for c in range(2, 9):
        ws2.cell(row=r, column=c).border = BORDER
        if i % 2 == 1:
            ws2.cell(row=r, column=c).fill = BAND_FILL

note_row = start_r + 1 + len(MONTH_ORDER) + 1
ws2.cell(row=note_row, column=2, value='หมายเหตุ: "จำนวนพนักงานที่บันทึก" นับจากรายชื่อไม่ซ้ำในข้อมูลดิบต่อเดือน (คำนวณล่วงหน้าด้วย Python เนื่องจากการนับ Distinct Count ต่อเงื่อนไขไม่รองรับสูตรมาตรฐานใน LibreOffice) ส่วน "ชั่วโมงที่บันทึกจริง" และ "อัตราการบันทึก" เป็นสูตรที่ดึงจาก Sheet ข้อมูลดิบโดยตรง').font = Font(name=FONT_NAME, size=9, italic=True, color='898781')
ws2.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=8)
ws2.row_dimensions[note_row].height = 30
autofit(ws2, [3, 10, 8, 22, 16, 16, 16, 16])

# ---------------------------------------------------------------
# Sheet 3: สรุปรายแผนก
# ---------------------------------------------------------------
ws3 = wb.create_sheet('สรุปรายแผนก')
ws3.sheet_view.showGridLines = False
ws3['B2'] = 'สรุปชั่วโมงทำงานตามฝ่าย (ม.ค.–ก.ค. 2569)'
ws3['B2'].font = TITLE_FONT

master_sheet_name = 'Master พนักงาน + Job Role'
dept_list = sorted([d for d in master_full['ชื่อหน่วยงาน (ระดับ 4)'].dropna().unique()])
headers3 = ['ฝ่าย', 'จำนวนพนักงาน (Master)', 'ชั่วโมงทำงานรวม (จาก Timesheet)']
start_r3 = 4
for j, h in enumerate(headers3):
    ws3.cell(row=start_r3, column=2 + j, value=h)
style_header_row(ws3, start_r3, len(headers3), start_col=2)

master_last_row = len(master_full) + 1
for i, dept in enumerate(dept_list):
    r = start_r3 + 1 + i
    ws3.cell(row=r, column=2, value=dept).font = BODY_FONT
    ws3.cell(row=r, column=3, value=(
        f"=COUNTIF('{master_sheet_name}'!$F$2:$F${master_last_row},B{r})"
    )).font = BODY_FONT
    ws3.cell(row=r, column=4, value=(
        f"=SUMIFS('{raw_sheet_name}'!$H$2:$H${raw_last_row},"
        f"'{raw_sheet_name}'!$J$2:$J${raw_last_row},B{r})"
    )).font = BODY_FONT
    for c in range(2, 5):
        ws3.cell(row=r, column=c).border = BORDER
        if i % 2 == 1:
            ws3.cell(row=r, column=c).fill = BAND_FILL
autofit(ws3, [3, 46, 20, 26])

# ---------------------------------------------------------------
# Sheet 4: Job Role Mapping Quality
# ---------------------------------------------------------------
ws4 = wb.create_sheet('Job Role Mapping Quality')
ws4.sheet_view.showGridLines = False
ws4['B2'] = 'คุณภาพการจับคู่ Job Role กับ PCC Job Role Database (พนักงาน PDE)'
ws4['B2'].font = TITLE_FONT
statuses = ['สูง (High)', 'ปานกลาง (Medium) - ควรตรวจสอบ', 'ไม่มี Job Code รองรับ']
headers4 = ['สถานะการแมส Job Role', 'จำนวนพนักงาน', '% ของพนักงานทั้งหมด']
start_r4 = 4
for j, h in enumerate(headers4):
    ws4.cell(row=start_r4, column=2 + j, value=h)
style_header_row(ws4, start_r4, len(headers4), start_col=2)
for i, s in enumerate(statuses):
    r = start_r4 + 1 + i
    ws4.cell(row=r, column=2, value=s).font = BODY_FONT
    ws4.cell(row=r, column=3, value=f"=COUNTIF('{master_sheet_name}'!$L$2:$L${master_last_row},B{r})").font = BODY_FONT
    ws4.cell(row=r, column=4, value=f"=C{r}/{master_last_row - 1}").font = BODY_FONT
    ws4.cell(row=r, column=4).number_format = '0.0%'
    for c in range(2, 5):
        ws4.cell(row=r, column=c).border = BORDER
        if i % 2 == 1:
            ws4.cell(row=r, column=c).fill = BAND_FILL
total_r = start_r4 + 1 + len(statuses)
ws4.cell(row=total_r, column=2, value='รวมพนักงาน PDE (Master)').font = BOLD_FONT
ws4.cell(row=total_r, column=3, value=master_last_row - 1).font = BOLD_FONT
for c in (2, 3, 4):
    ws4.cell(row=total_r, column=c).border = BORDER

note_r = total_r + 2
ws4.cell(row=note_r, column=2, value='"ไม่มี Job Code รองรับ" หมายถึง Job Role เดิมของพนักงานยังไม่ถูกจับคู่กับตำแหน่งใน PCC Job Role Database ทำให้ยังไม่มี KPI/JE มาตรฐานสำหรับเทียบ Performance ในโครงการนี้ ควรจัดลำดับความสำคัญให้ตรวจสอบกลุ่มนี้ก่อนเริ่ม Phase 1').font = Font(name=FONT_NAME, size=9, italic=True, color='898781')
ws4.merge_cells(start_row=note_r, start_column=2, end_row=note_r, end_column=6)
ws4.row_dimensions[note_r].height = 30
autofit(ws4, [3, 34, 16, 18])

# ---------------------------------------------------------------
# Sheet 5: สรุปรายบุคคล
# ---------------------------------------------------------------
ws5 = wb.create_sheet('สรุปรายบุคคล')
ws5.sheet_view.showGridLines = False
ws5['B2'] = 'สรุปชั่วโมงทำงานและสัดส่วนกิจกรรมรายบุคคล (ม.ค.–ก.ค. 2569)'
ws5['B2'].font = TITLE_FONT

cats = ['Project / Job code', 'Project work', 'Admin / งานธุรการ', 'IT Support (ticket)', 'Leave / ลา', 'Other / ไม่ระบุหมวด']
headers5 = ['No.', 'ชื่อ-สกุล', 'ฝ่าย', 'Job Role', 'สถานะการแมส Job Role', 'ชั่วโมงรวม'] + [f'{c} (ชม.)' for c in cats] + ['จำนวน Job/โครงการที่ต่างกัน', 'เดือนที่มีการบันทึก']
start_r5 = 4
for j, h in enumerate(headers5):
    ws5.cell(row=start_r5, column=2 + j, value=h)
style_header_row(ws5, start_r5, len(headers5), start_col=2)

emp_sorted = emp_master_df.sort_values('full_name').reset_index(drop=True)
for i, row in emp_sorted.iterrows():
    r = start_r5 + 1 + i
    name = row['full_name']
    ws5.cell(row=r, column=2, value=i + 1).font = BODY_FONT
    ws5.cell(row=r, column=3, value=name).font = BODY_FONT
    ws5.cell(row=r, column=4, value=(
        f"=IFERROR(INDEX('{master_sheet_name}'!$F$2:$F${master_last_row},MATCH(C{r},'{master_sheet_name}'!$B$2:$B${master_last_row},0)),\"-\")"
    )).font = BODY_FONT
    ws5.cell(row=r, column=5, value=(
        f"=IFERROR(INDEX('{master_sheet_name}'!$C$2:$C${master_last_row},MATCH(C{r},'{master_sheet_name}'!$B$2:$B${master_last_row},0)),\"-\")"
    )).font = BODY_FONT
    ws5.cell(row=r, column=6, value=(
        f"=IFERROR(INDEX('{master_sheet_name}'!$L$2:$L${master_last_row},MATCH(C{r},'{master_sheet_name}'!$B$2:$B${master_last_row},0)),\"-\")"
    )).font = BODY_FONT
    total_col_letter_start = 7
    ws5.cell(row=r, column=total_col_letter_start, value=(
        f"=SUMIF('{raw_sheet_name}'!$A$2:$A${raw_last_row},C{r},'{raw_sheet_name}'!$H$2:$H${raw_last_row})"
    )).font = BOLD_FONT
    for k, cat in enumerate(cats):
        col = total_col_letter_start + 1 + k
        ws5.cell(row=r, column=col, value=(
            f"=SUMIFS('{raw_sheet_name}'!$H$2:$H${raw_last_row},"
            f"'{raw_sheet_name}'!$A$2:$A${raw_last_row},C{r},"
            f"'{raw_sheet_name}'!$I$2:$I${raw_last_row},\"{cat}\")"
        )).font = BODY_FONT
    dj = int(distinct_jobs.get(name, 0))
    ma = int(months_active.get(name, 0))
    col_dj = total_col_letter_start + 1 + len(cats)
    ws5.cell(row=r, column=col_dj, value=dj).font = BODY_FONT
    ws5.cell(row=r, column=col_dj + 1, value=ma).font = BODY_FONT
    for c in range(2, col_dj + 2):
        ws5.cell(row=r, column=c).border = BORDER
        if i % 2 == 1:
            ws5.cell(row=r, column=c).fill = BAND_FILL

note_r5 = start_r5 + 1 + len(emp_sorted) + 1
ws5.cell(row=note_r5, column=2, value='คอลัมน์ "จำนวน Job/โครงการที่ต่างกัน" และ "เดือนที่มีการบันทึก" คำนวณล่วงหน้าด้วย Python (Distinct Count) เนื่องจาก LibreOffice ไม่รองรับสูตรนับค่าไม่ซ้ำแบบมีเงื่อนไขได้เสถียร คอลัมน์อื่นในชีตนี้เป็นสูตรที่ดึงจาก Sheet ข้อมูลดิบและ Master โดยตรง').font = Font(name=FONT_NAME, size=9, italic=True, color='898781')
ws5.merge_cells(start_row=note_r5, start_column=2, end_row=note_r5, end_column=8)
ws5.row_dimensions[note_r5].height = 30
autofit(ws5, [3, 5, 24, 42, 34, 26, 11] + [16] * len(cats) + [16, 14])

# ---------------------------------------------------------------
# Sheet 6: Master พนักงาน + Job Role  (reference data — plain values)
# ---------------------------------------------------------------
ws6 = wb.create_sheet(master_sheet_name)
ws6.sheet_view.showGridLines = False
cols6 = ['Employee Code', 'full_name', 'Job Role', 'Position', 'Employee Level', 'Department',
         'Job Code (PCC)', 'Job Role DB (TH)', 'Job Family', 'JE รวม', 'Match %', 'สถานะการแมส',
         'Job Role Code', 'มี Job Role Code ซ้ำกับพนักงานอื่นที่แมสต่างกัน']
# NOTE: columns A-L keep the exact order/letters relied on by INDEX/MATCH & COUNTIF/SUMIF
# formulas in other sheets (สรุปรายบุคคล, สรุปรายแผนก, Job Role Mapping Quality, ข้อมูลดิบ
# Timesheet) — the two new columns from the Job-Role-Code join are appended as M, N so no
# existing formula reference shifts.
df6 = emp_master_df[cols6].rename(columns={'full_name': 'ชื่อ-สกุล'})
write_df(ws6, df6, start_row=1, number_formats={'Match %': '0.0"%"'})
autofit(ws6, [16, 24, 40, 30, 14, 42, 16, 32, 26, 10, 12, 30, 18, 34])

# ---------------------------------------------------------------
# Sheet 7: ข้อมูลดิบ Timesheet (raw + derived category column)
# ---------------------------------------------------------------
ws7 = wb.create_sheet(raw_sheet_name)
ws7.sheet_view.showGridLines = False
cols7 = ['employee_name', 'year', 'month', 'month_name', 'date_str', 'job_no', 'description', 'hours', 'category']
df7 = ts[cols7].rename(columns={
    'employee_name': 'ชื่อ-สกุล', 'year': 'ปี', 'month': 'เดือน (เลข)', 'month_name': 'เดือน',
    'date_str': 'วันที่', 'job_no': 'Job No.', 'description': 'Description', 'hours': 'ชั่วโมง',
    'category': 'หมวดกิจกรรม (จัดกลุ่มอัตโนมัติ)',
})
last_row7 = write_df(ws7, df7, start_row=1)
ws7.cell(row=1, column=10, value='ฝ่าย (lookup จาก Master)')
ws7.cell(row=1, column=10).font = HEADER_FONT
ws7.cell(row=1, column=10).fill = HEADER_FILL
ws7.cell(row=1, column=10).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws7.cell(row=1, column=10).border = BORDER
for i in range(len(df7)):
    r = 2 + i
    cell = ws7.cell(row=r, column=10, value=(
        f"=IFERROR(INDEX('{master_sheet_name}'!$F$2:$F${master_last_row},MATCH(A{r},'{master_sheet_name}'!$B$2:$B${master_last_row},0)),\"-\")"
    ))
    cell.font = BODY_FONT
    cell.border = BORDER
    if i % 2 == 1:
        cell.fill = BAND_FILL
autofit(ws7, [24, 8, 12, 10, 10, 22, 60, 10, 26, 42])
ws7['L1'] = 'หมวดกิจกรรมจัดกลุ่มจาก Job No./Description ด้วยกฎคำสำคัญ (ดูรายละเอียดใน Sheet วิธีใช้งาน)'
ws7['L1'].font = Font(name=FONT_NAME, size=9, italic=True, color='898781')

# reorder sheets
order = ['วิธีใช้งาน', 'สรุปภาพรวมรายเดือน', 'สรุปรายแผนก', 'Job Role Mapping Quality',
         'สรุปรายบุคคล', master_sheet_name, raw_sheet_name]
wb._sheets = [wb[name] for name in order]
for name in order:
    wb[name].sheet_view.showGridLines = False

wb.save('/tmp/analysis/PDE_วิเคราะห์ข้อมูลพนักงาน.xlsx')
print('saved')
